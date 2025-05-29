
from flask import Flask, request, render_template_string, send_file
from werkzeug.utils import secure_filename
import openpyxl
import tempfile
import os
from PIL import Image, ImageDraw, ImageFont
from barcode import Code128
from barcode.writer import ImageWriter

app = Flask(__name__)

HTML_FORM = '''
<!DOCTYPE html>
<html>
<head>
    <title>IMEI Label Printer</title>
</head>
<body>
    <h2>Upload Excel File</h2>
    <form method="post" enctype="multipart/form-data">
        <input type="file" name="excel"><br><br>
        <input type="text" name="imei" placeholder="Enter IMEI"><br><br>
        <input type="submit" value="Generate Label"><br><br>
    </form>
    <form method="post" enctype="multipart/form-data" action="/final">
        <input type="file" name="excel_final"><br><br>
        <input type="text" name="imei_final" placeholder="Enter IMEI for Final Label"><br><br>
        <input type="submit" value="Print Label Final"><br><br>
    </form>
    <button onclick="window.print()">🖨️ Print This Page</button>
</body>
</html>
'''

def create_label_image(item):
    width, height = 696, 1200
    img = Image.new('RGB', (width, height), 'white')
    draw = ImageDraw.Draw(img)

    try:
        font = ImageFont.truetype("arial.ttf", 22)
        font_bold = ImageFont.truetype("arialbd.ttf", 26)
    except:
        font = ImageFont.load_default()
        font_bold = ImageFont.load_default()

    y = 20
    barcode_val = item.get("IMEI") or item.get("ID")
    if barcode_val:
        try:
            barcode = Code128(str(barcode_val), writer=ImageWriter())
            with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_bar:
                barcode_path = tmp_bar.name
                barcode.save(barcode_path[:-4])
                barcode_img = Image.open(barcode_path)
                barcode_img = barcode_img.resize((500, 100))
                img.paste(barcode_img, (int((width - 500) / 2), y))
                y += 110
        except Exception as e:
            draw.text((30, y), f"Barcode error: {e}", font=font, fill="red")
            y += 30

    for key, value in item.items():
        if not value:
            continue
        draw.text((30, y), f"{key}:", font=font_bold, fill="black")
        y += 30

        max_width = 600
        words = value.split()
        line = ""
        for word in words:
            test_line = f"{line} {word}".strip()
            text_width = draw.textbbox((0, 0), test_line, font=font)[2]
            if text_width <= max_width:
                line = test_line
            else:
                draw.text((60, y), line, font=font, fill="black")
                y += 28
                line = word
        if line:
            draw.text((60, y), line, font=font, fill="black")
            y += 35
        y += 5

    return img

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        file = request.files['excel']
        imei_input = request.form.get('imei', '').strip()
        if not file or not imei_input:
            return "Please upload a file and enter IMEI."

        filename = secure_filename(file.filename)
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            file.save(tmp.name)
            wb = openpyxl.load_workbook(tmp.name, data_only=True)
            sheet = wb.active
            headers = [str(cell.value).strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = {headers[i]: str(cell).strip() if cell else "" for i, cell in enumerate(row)}
                if imei_input in row_data.values():
                    label_img = create_label_image(row_data)
                    label_path = os.path.join(tempfile.gettempdir(), "label_output.png")
                    label_img.save(label_path)
                    
    with open(label_path, "rb") as f:
        image_data = f.read()
    
import base64
encoded = base64.b64encode(image_data).decode("utf-8")

    return f'''
    <html>
    <head><title>Print Label</title></head>
    <body onload="window.print()">
        <img src="data:image/png;base64,{encoded}" width="600"/>
    </body>
    </html>
    '''
    
        return "IMEI not found in uploaded Excel."

    return render_template_string(HTML_FORM)

@app.route('/final', methods=['POST'])
def print_final_label():
    file = request.files.get('excel_final')
    imei_input = request.form.get('imei_final')
    if not file or not imei_input:
        return "Missing input."

    filename = secure_filename(file.filename)
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        file.save(tmp.name)
        wb = openpyxl.load_workbook(tmp.name, data_only=True)
        sheet = wb.active
        headers = [str(cell.value).strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {headers[i]: str(cell).strip() if cell else "" for i, cell in enumerate(row)}
            if imei_input in row_data.values():
                img = Image.new("RGB", (696, 400), "white")
                draw = ImageDraw.Draw(img)
                try:
                    font_bold = ImageFont.truetype("arialbd.ttf", 28)
                    font = ImageFont.truetype("arial.ttf", 24)
                except:
                    font = ImageFont.load_default()
                    font_bold = ImageFont.load_default()

                y = 20
                draw.text((30, y), f"IMEI: {row_data.get('IMEI', '')}", font=font, fill="black")
                y += 40
                draw.text((30, y), f"ID: {row_data.get('ID', '')}", font=font, fill="black")
                y += 40
                draw.text((30, y), f"Carrier: {row_data.get('Carrier', '')}", font=font, fill="black")
                y += 40
                draw.text((30, y), f"Status: {row_data.get('Status', '')}", font=font, fill="black")

                label_path = os.path.join(tempfile.gettempdir(), "final_label_output.png")
                img.save(label_path)
                
    with open(label_path, "rb") as f:
        image_data = f.read()
    
import base64
encoded = base64.b64encode(image_data).decode("utf-8")

    return f'''
    <html>
    <head><title>Print Label</title></head>
    <body onload="window.print()">
        <img src="data:image/png;base64,{encoded}" width="600"/>
    </body>
    </html>
    '''
    
    return "IMEI not found in Excel."

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')
